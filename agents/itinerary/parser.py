"""Parse itineraries from PDF and Excel files."""

from __future__ import annotations

import json
import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
import pdfplumber

from agents.common.llm import SONNET, make_llm

from .models import Itinerary, ItineraryItem, Location


def fix_json_string(json_str: str) -> str:
    """Fix common JSON issues that Claude sometimes produces."""
    # Remove trailing commas before ] or }
    json_str = re.sub(r",\s*([}\]])", r"\1", json_str)

    # Fix unescaped newlines inside strings (common issue)
    # This is tricky - we need to find strings and escape newlines within them
    # Simple approach: replace literal newlines that appear to be inside strings
    lines = json_str.split("\n")
    fixed_lines = []
    in_string = False
    for line in lines:
        # Count unescaped quotes to track if we're in a string
        quote_count = len(re.findall(r'(?<!\\)"', line))
        if in_string:
            # We're continuing a string from previous line - this is the problem
            # Escape it and continue
            line = "\\n" + line.replace("\n", "\\n").replace("\r", "\\r")
        fixed_lines.append(line)
        # Update string state (odd number of quotes toggles state)
        if quote_count % 2 == 1:
            in_string = not in_string

    json_str = "\n".join(fixed_lines)

    # Remove control characters except newlines and tabs
    json_str = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", json_str)

    return json_str


def _build_extraction_prompt() -> str:
    """Render the extraction prompt with today's date baked in.

    The year-resolution rule MUST be re-rendered each call because
    confirmations rarely include the year (e.g. "Wed, 20MAY"). Without
    today's date in the prompt the LLM picks whichever year matches the
    weekday, often 5+ years in the past. Use plain str.replace because
    the prompt body contains JSON example braces that would trip
    str.format.
    """
    today = datetime.now()
    return (
        _EXTRACTION_PROMPT_TEMPLATE.replace("{{CURRENT_DATE}}", today.strftime("%Y-%m-%d"))
        .replace("{{CURRENT_YEAR}}", str(today.year))
        .replace("{{NEXT_YEAR}}", str(today.year + 1))
    )


_EXTRACTION_PROMPT_TEMPLATE = """You are an expert at extracting structured travel itinerary data from text.

Today's date is {{CURRENT_DATE}}.

When a date in the source omits the year (very common on flight confirmations like "Wed, 20MAY"), use this rule:
- If the month/day is still upcoming this year, use {{CURRENT_YEAR}}
- If the month/day has already passed this year, use {{NEXT_YEAR}}
- Never pick a year more than 1 year in the past based on weekday matching alone

Analyze the following text from a travel itinerary document and extract all relevant information.

Return a JSON object with this exact structure:
{
    "title": "Trip title or destination",
    "destination_region": "The main destination region/state/country for context (e.g. 'Alaska, USA')",
    "start_date": "YYYY-MM-DD or null",
    "end_date": "YYYY-MM-DD or null",
    "travelers": ["list of traveler names"],
    "items": [
        {
            "title": "Activity/event title - for hotels, use the ACTUAL HOTEL NAME (e.g. 'Taj Palace' not 'Hotel stay')",
            "location_name": "Place name with region for geocoding (e.g. 'Seward, Alaska' not just 'Seward')",
            "location_address": "Full address if available",
            "location_type": "hotel|restaurant|attraction|airport|train_station|home|other",
            "date": "YYYY-MM-DD or null",
            "end_date": "YYYY-MM-DD or null (checkout date for hotels, return date for rentals)",
            "start_time": "HH:MM (24hr) or null",
            "end_time": "HH:MM (24hr) or null",
            "description": "Brief description",
            "category": "flight|train|bus|hotel|activity|meal|transport|home|other",
            "confirmation_number": "if available or null",
            "notes": "any additional notes",
            "day_number": day number as integer or null,
            "is_home_location": true ONLY for the very first outbound flight/leg leaving the traveler's home city - NOT for the return flight home
        }
    ]
}

Important:
- Extract ALL items/events from the itinerary
- Be precise with dates and times
- ALWAYS include the state/region/country with location names for accurate geocoding (e.g. "Homer, Alaska" not just "Homer")
- Set is_home_location: true ONLY for the very first departing flight/train (the one leaving the traveler's home city to start the trip). The return flight home at the end of the trip should have is_home_location: false - it is a real trip event. All destination activities, hotels, and connecting flights should have is_home_location: false.
- For HOTELS/ACCOMMODATIONS: Extract the ACTUAL hotel name and set end_date to the checkout date if available. (e.g. "Taj Palace", "ITC Mughal", "Marriott") NOT generic descriptions like "Hotel stay in Delhi". The hotel name should go in the "title" field.
- For FLIGHTS: start_time is DEPARTURE time, end_time is ARRIVAL time. Include both if available. Title should include flight number and route (e.g. "UA 123 SFO → JFK"). Keep IATA airport codes as-is in the title - do NOT try to expand them to city names (e.g. keep "DEN → BIH" not "Denver → Birmingham"). The location_name should be the destination airport code only (e.g. "BIH" not "Birmingham")
- For TRAINS (AVE, TGV, Eurostar, Amtrak, subway, metro, rail, etc.): set category to "train". start_time is DEPARTURE time, end_time is ARRIVAL time. Include both if available.
- For BUSES (coach, intercity bus): set category to "bus".
- Use "transport" only for car rentals, taxis, and transfers that are not trains or buses.
- For MEALS/RESERVATIONS: start_time is reservation time, end_time can be estimated end (e.g. +2 hours for dinner)
- If information is not available, use null
- Ensure the JSON is valid

CRITICAL for day_number:
- You MUST set day_number for ALL items that have a date or are associated with a specific day
- If the document shows "Day 1", "Day 2", etc., use those numbers directly
- If the document shows dates like "Jan 15", "Jan 16", calculate day_number as: Day 1 = first date, Day 2 = second date, etc.
- If an item is listed under a specific date heading (e.g., "Saturday, January 15"), it MUST have a day_number
- Only set day_number to null for items that are truly unscheduled (e.g., "Ideas", "To Do Later", "Recommendations" sections)
- Items listed under date headings or "Day X" headings should NEVER have null day_number

Text to analyze:
"""


class ItineraryParser:
    """Parse itineraries from various file formats using Claude."""

    def __init__(self, api_key: str | None = None):
        # api_key param kept for backwards compatibility; fla reads ANTHROPIC_API_KEY from env
        self.llm = make_llm(model=SONNET, max_tokens=8192)

    def parse_file(self, file_path: str | Path) -> Itinerary:
        """Parse an itinerary from a PDF or Excel file."""
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix == ".pdf":
            text = self._extract_text_from_pdf(file_path)
        elif suffix in (".xlsx", ".xls"):
            text = self._extract_text_from_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

        return self._parse_with_claude(text, str(file_path))

    def parse_text(self, text: str, source_url: str = "unknown") -> Itinerary:
        """Parse an itinerary from raw text (e.g., from HTML page)."""
        return self._parse_with_claude(text, source_url)

    def parse_image(self, image_data: str, media_type: str, source_file: str) -> Itinerary:
        """Parse an itinerary from a base64-encoded image (PNG/JPG/etc.).

        Calls Claude's vision API with the image plus the standard extraction
        prompt. Used for screenshots and image confirmations that the
        text-based parse_file() can't handle.
        """
        messages = [
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        },
                    },
                    {"type": "text", "text": _build_extraction_prompt()},
                ],
            }
        ]
        response_text = self.llm.call_api(system_prompt="", messages=messages)
        return self._parse_response_text(response_text, source_file)

    def _parse_response_text(self, response_text: str, source_file: str) -> Itinerary:
        """Shared post-processing: pull JSON out of an LLM response and build
        an Itinerary. Factored out so parse_text and parse_image can share it.
        """
        if "```json" in response_text:
            json_str = response_text.split("```json")[1].split("```")[0].strip()
        elif "```" in response_text:
            json_str = response_text.split("```")[1].split("```")[0].strip()
        else:
            json_str = response_text.strip()
        json_str = fix_json_string(json_str)
        try:
            data = json.loads(json_str)
        except json.JSONDecodeError as e:
            lines = json_str.split("\n")
            error_line = e.lineno - 1 if e.lineno else 0
            start = max(0, error_line - 2)
            end = min(len(lines), error_line + 3)
            context = "\n".join(f"{i + 1}: {lines[i]}" for i in range(start, end))
            raise ValueError(
                f"Failed to parse Claude's response as JSON: {e}\nContext:\n{context}"
            ) from e
        return self._build_itinerary(data, source_file)

    def _extract_text_from_pdf(self, file_path: Path) -> str:
        """Extract text content from a PDF file."""
        text_parts = []

        # Try pdfplumber first (better table extraction)
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text_parts.append(page_text)

                        # Also extract tables
                        tables = page.extract_tables()
                        for table in tables:
                            for row in table:
                                if row:
                                    row_text = " | ".join(str(cell) for cell in row if cell)
                                    text_parts.append(row_text)
                    except Exception as e:
                        # Some pages may have malformed color values or other issues
                        print(f"Warning: pdfplumber could not extract page: {e}")
                        continue
        except Exception as e:
            print(f"Warning: pdfplumber failed: {e}")

        # If pdfplumber failed or got no text, try PyPDF2 as fallback
        if not text_parts:
            print("Trying PyPDF2 as fallback...")
            try:
                from PyPDF2 import PdfReader

                reader = PdfReader(file_path)
                for page in reader.pages:
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text_parts.append(page_text)
                    except Exception as e:
                        print(f"Warning: PyPDF2 could not extract page: {e}")
                        continue
            except Exception as e:
                print(f"Warning: PyPDF2 also failed: {e}")

        if not text_parts:
            raise ValueError(
                "Could not extract any text from PDF. The file may be image-based or corrupted."
            )

        return "\n\n".join(text_parts)

    def _extract_text_from_excel(self, file_path: Path) -> str:
        """Extract text content from an Excel file."""
        text_parts = []
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== Sheet: {sheet_name} ===")

            for row in sheet.iter_rows():
                row_values = []
                for cell in row:
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                if row_values:
                    text_parts.append(" | ".join(row_values))

        return "\n".join(text_parts)

    def _parse_with_claude(self, text: str, source_file: str) -> Itinerary:
        """Use Claude to extract structured data from itinerary text."""
        response_text = self.llm.call_api(
            system_prompt="",
            messages=[{"role": "user", "content": _build_extraction_prompt() + text}],
        )
        return self._parse_response_text(response_text, source_file)

    def _build_itinerary(self, data: dict, source_file: str) -> Itinerary:
        """Build an Itinerary object from parsed data."""
        items = []

        for item_data in data.get("items", []):
            location = Location(
                name=item_data.get("location_name", "Unknown"),
                address=item_data.get("location_address"),
                location_type=item_data.get("location_type"),
            )

            item = ItineraryItem(
                title=item_data.get("title", "Untitled"),
                location=location,
                date=self._parse_date(item_data.get("date")),
                end_date=self._parse_date(item_data.get("end_date")),
                start_time=self._parse_time(item_data.get("start_time")),
                end_time=self._parse_time(item_data.get("end_time")),
                description=item_data.get("description"),
                category=item_data.get("category"),
                confirmation_number=item_data.get("confirmation_number"),
                notes=item_data.get("notes"),
                day_number=item_data.get("day_number"),
                is_home_location=item_data.get("is_home_location", False),
            )
            items.append(item)

        return Itinerary(
            title=data.get("title", "Untitled Itinerary"),
            items=items,
            start_date=self._parse_date(data.get("start_date")),
            end_date=self._parse_date(data.get("end_date")),
            travelers=data.get("travelers", []),
            source_file=source_file,
        )

    def _parse_date(self, date_str: str | None) -> date | None:
        """Parse a date string to a date object."""
        if not date_str:
            return None
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return None

    def _parse_time(self, time_str: str | None) -> time | None:
        """Parse a time string to a time object."""
        if not time_str:
            return None
        try:
            return datetime.strptime(time_str, "%H:%M").time()
        except ValueError:
            return None
